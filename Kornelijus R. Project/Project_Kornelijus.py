import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from geopy.geocoders import Nominatim
import folium

def find_restaurant():    
    # === Step 1: Collect user input ===
    restaurant = input("Restaurant name (include city if needed): ")
    date_visited = input("Date visited (YYYY-MM-DD): ")
    notes = input("Meeting notes: ")

    # === Step 2: Excel setup ===
    file_name = "restaurant_meetings.xlsx"
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Meetings"
        ws.append(["Restaurant", "Date", "Address", "Notes", "Latitude", "Longitude"])
    else:
        wb = load_workbook(file_name)
        ws = wb.active

    # === Step 3: Geocode using restaurant name ===
    from geopy.geocoders import Nominatim

    geolocator = Nominatim(user_agent="restaurant_logger")

    # Try automatic geocoding
    location = geolocator.geocode(restaurant)

    if not location:
        print("⚠️ Could not find the restaurant automatically.")
        manual_address = input("Please enter the full address manually: ")
        location = geolocator.geocode(manual_address)
        if location:
            full_address = location.address
            lat, lon = location.latitude, location.longitude
            print(f"✅ Found manually: {full_address} ({lat}, {lon})")
        else:
            full_address = "Not Found"
            lat, lon = None, None
            print("❌ Still couldn't find the location.")
    else:
        full_address = location.address
        lat, lon = location.latitude, location.longitude
        print(f"✅ Found: {full_address} ({lat}, {lon})")


    # === Step 4: Save to Excel ===
    ws.append([restaurant, date_visited, full_address, notes, lat, lon])
    wb.save(file_name)

    # === Step 5: Create map with ALL pins ===
    map_file = "restaurant_map.html"

    if lat and lon:
        m = folium.Map(location=[lat, lon], zoom_start=12)
    else:
        m = folium.Map(location=[0, 0], zoom_start=2)

    # Add all markers from Excel
    for row in ws.iter_rows(min_row=2, values_only=True):
        r_name, r_date, r_address, r_notes, r_lat, r_lon = row
        if r_lat and r_lon:
            folium.Marker(
                [r_lat, r_lon],
                popup=f"{r_name} ({r_date})\n{r_notes}",
                tooltip=r_name
            ).add_to(m)

    m.save(map_file)
    print("✅ Map updated with all restaurants.")

    import webbrowser
    import subprocess
    import platform

    # Open Excel file
    try:
        if platform.system() == "Windows":
            os.startfile(file_name)
        elif platform.system() == "Darwin":  # macOS
            subprocess.call(["open", file_name])
        else:
            subprocess.call(["xdg-open", file_name])  # Linux
    except Exception as e:
        print(f"Could not open Excel file: {e}")

    # Open map in browser (cross-platform)
    try:
        webbrowser.open(map_file)
    except Exception as e:
        print(f"Could not open map: {e}")


if __name__ == "__main__":
    find_restaurant()